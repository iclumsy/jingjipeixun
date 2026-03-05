"""
数据导出路由。

本模块提供学员数据的 Excel 导出功能，用于管理员批量导出学员信息。

API 端点:
    GET /api/export/excel - 导出学员数据为 Excel 文件

导出特性:
    - 支持按状态、公司、培训类型筛选导出范围
    - Excel 表头加粗居中，数据行带边框
    - 自动调整列宽以适应内容
    - 文件名格式: 学员信息_YYYYMMDD_HHMMSS.xlsx
"""
from flask import Blueprint, request, jsonify, send_file, current_app
from models.student import get_students
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 创建导出蓝图
export_bp = Blueprint('export', __name__)


@export_bp.route('/api/export/excel', methods=['GET'])
def export_excel():
    """
    将学员数据导出为 Excel 文件。

    使用 openpyxl 库在内存中生成 Excel 工作簿。
    审核状态会从英文转换为中文显示。

    查询参数:
        status (str)        : 按审核状态筛选
        company (str)       : 按公司名称筛选
        training_type (str) : 按培训类型筛选

    返回:
        200: Excel 文件流
        500: 导出失败
    """
    try:
        # 获取筛选参数
        status = request.args.get('status', '')
        company = request.args.get('company', '')

        # 根据筛选条件查询学员数据
        training_type = request.args.get('training_type', '')
        students = get_students(status, '', company, training_type)

        # 创建 Excel 工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "学员信息"

        # 定义 Excel 表头（与数据库字段一一对应）
        headers = [
            'ID', '姓名', '性别', '文化程度', '毕业院校', '所学专业',
            '身份证号', '手机号', '单位名称', '单位地址',
            '作业类别', '操作项目', '项目代号',
            '状态', '创建时间'
        ]

        # 写入表头行（第一行），设置加粗、居中、边框样式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 逐行写入学员数据（从第2行开始）
        for row_num, student in enumerate(students, 2):
            # 将数据库状态值转换为中文显示
            status_text = '已审核' if student['status'] == 'reviewed' else '未审核'

            data = [
                student['id'],
                student['name'],
                student['gender'],
                student['education'],
                student.get('school', ''),
                student.get('major', ''),
                student['id_card'],
                student['phone'],
                student.get('company', ''),
                student.get('company_address', ''),
                student['job_category'],
                student.get('exam_project', ''),
                student.get('project_code', ''),
                status_text,
                student.get('created_at', '')
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # 自动调整列宽：取每列最大内容长度 + 2 字符余量，上限 50
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 将工作簿保存到内存缓冲区（避免写入临时文件）
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        current_app.logger.info(f'Excel export generated with {len(students)} students')

        # 返回 Excel 文件作为附件下载
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'学员信息_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        current_app.logger.error(f'Error exporting to Excel: {str(e)}')
        return jsonify({'error': str(e)}), 500
